import pandas as pd
import numpy
from league_standings import *
from openpyxl import load_workbook
import time


html_header = '''<!DOCTYPE html>
<html>
<head>
<style>
#brackets {
  font-family: Arial, Helvetica, sans-serif;
  border-collapse: collapse;
  width: 100%;
}

#brackets td, #brackets th {
  border: 1px solid #ddd;
  padding: 8px;
}

#brackets tr:nth-child(even){background-color: #f2f2f2;}

#brackets tr:hover {background-color: #ddd;}

#brackets th {
  padding-top: 12px;
  padding-bottom: 12px;
  text-align: left;
  background-color: #357330;
  color: white;
}
</style>
</head>
<body>'''

class BracketServer:

    def __init__(self, spreadsheet_path):
        self.spreadsheet_path = spreadsheet_path
        self.index_file = r'index.html'
        self.users = [["Mark", "D"], ["Ush", "I"], ["Joey", "N"], ["Ru", "S"], ["Dan", "X"]]
        self.east_map = {}
        self.west_map = {}

    def updateStandings(self):
        # get top 8 nba teams of each conference and write to spreadsheet

        east, west = getStandings()

        for idx, team in enumerate(east):
            self.east_map[team] = idx + 1
        for idx, team in enumerate(west):
            self.west_map[team] = idx + 1

        book = load_workbook(self.spreadsheet_path)
        sheet = book.active

        for i in range(8):
            sheet.cell(column=1, row=i+4, value=east[i])
        for i in range(8):
            sheet.cell(column=1, row=i+15, value=west[i])

        book.save(self.spreadsheet_path)

    def updateHTML(self):
        # write html to file
        f = open(self.index_file, "w")
        f.write(html_header)
        f.write(self._convertSheetToHTML())
        f.write('</html>')
        f.close()
        
    def _convertSheetToHTML(self):
        # create html from data in spreadsheet

        scores = {}

        # current east standings
        curr_standings = pd.read_excel(self.spreadsheet_path, usecols="A")
        curr_standings_east = curr_standings.iloc[2:10]
        curr_standings_east = curr_standings_east.rename(columns={'INPUT NAME>': 'East'})
        # current west standings
        curr_standings_west = curr_standings.iloc[13:21]
        curr_standings_west = curr_standings_west.rename(columns={'INPUT NAME>': 'West'})
        # combine east and west into one dataframe
        curr_standings = pd.concat([curr_standings_west.reset_index(drop=True), curr_standings_east.reset_index(drop=True)], axis=1)
        curr_standings.index = range(1, 9)
        html_string = "<h1>Current Standings</h1>" + self._setTableID(curr_standings.to_html())

        # user predictions and scores
        for user in self.users:

            user_name = user[0]
            user_col = user[1]

            html_string = html_string + "<h1> " + user[0] + " </h1>"
            # east
            user_df = pd.read_excel(self.spreadsheet_path, usecols=user_col)
            user_east = user_df.iloc[2:10]
            user_east = user_east.rename(columns={user_name: user_name + "-East"})
            user_east["Position"] = [self.east_map[x] if x in self.east_map else None for x in user_east[user_name + "-East"]]
            user_east["Points"] = [8 - abs(idx + 1 - x) for idx, x in enumerate(user_east["Position"])]
            # west
            user_west = user_df.iloc[13:21]
            user_west = user_west.rename(columns={user_name: user_name + "-West"})
            user_west["Position"] = [self.west_map[x] if x in self.west_map else None for x in user_west[user_name + "-West"]]
            user_west["Points"] = [8 - abs(idx + 1 - x) for idx, x in enumerate(user_west["Position"])]
            # combined
            user_df = pd.concat([user_west.reset_index(drop=True), user_east.reset_index(drop=True)], axis=1)
            user_df.index = range(1, 9)
            scores[user_name] = str(int(numpy.nansum(user_west["Points"]) + numpy.nansum(user_east["Points"])))
            html_string = html_string + self._setTableID(user_df.to_html()) + "<h2> " + user_name + " Score: " + scores[user_name] + "</h2>"

        scores = {k: v for k, v in sorted(scores.items(), key=lambda item: item[1], reverse=True)}
        
        # add table with everyone's score
        d = {'name': scores.keys(), 'score': scores.values()}
        score_df = pd.DataFrame(data=d)
        score_df.index = range(1, 6)
        html_string = "<h1>Ranking</h1>" + self._setTableID(score_df.to_html()) + html_string
       
        return html_string.replace("NaN", "")

    def _setTableID(self, html_string):
        return html_string[0:7] + 'id="brackets" ' + html_string[8:]


bracketSvr = BracketServer(r'NBA Playoffs Scoring.xlsx')

while True:
    bracketSvr.updateStandings()
    bracketSvr.updateHTML()
    print("updated")
    time.sleep(60)
